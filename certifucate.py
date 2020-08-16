import cv2 as cv
import openpyxl


temp_path="O1.png"
detail_path=r"Recruitment.xlsx"
output=r"C:\Users\Ankit Giri\Desktop\New folder"
font_size=4

font_color=(0,0,0)
font_color1=(128,128,128)
x=-50
y=-250

obj=openpyxl.load_workbook(detail_path)
sheet=obj.active
for i in range(2,81):
    get_name=sheet.cell(row=i,column=3)
    ceti_name=get_name.value
    
    get_name1=sheet.cell(row=i,column=7)
    ceti_name1=get_name1.value
    
    img=cv.imread(temp_path)
    font=cv.FONT_HERSHEY_DUPLEX
    font1=cv.FONT_HERSHEY_SIMPLEX
    text_size=cv.getTextSize(ceti_name, font , font_size, 10)[0]
    text_x=(img.shape[1]-text_size[0])/2 + x
    text_y=(img.shape[0]+text_size[1])/2-y
    text_x=int(text_x)
    text_y=int(text_y)
    cv.putText(img,ceti_name,(text_x,text_y),font,font_size,font_color,10)

   
    certi_path=ceti_name1+".png"
    cv.imwrite(certi_path,img)

    

